from __future__ import annotations

from copy import copy
from io import BytesIO
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from src.motor import ResultadoMotor


COLUNAS_BASE = [
    "Loja", "Ciclo", "Tratativa", "Fabricante ", "Categoria ", "Condição Pagamento",
    "Condição Leilão", "Código Fornecedor", "Prazo Pagamento", "SKU", "EAN",
    "Tipo de Produto", "Descrição", "Quantidade Solicitada", "Variação 1", "Status ",
    "Novo Fornecedor ", "Novo Código Fornecedor ", "Nova Variação", "Unidade", "Preço",
    "Prazo Entrega", "Embalagem", "Tipo Condição", "ComercialDiscount",
    "FinancialDiscount", "E-Mail", "NR COTAÇÃO", "Comprador ",
]

COLUNAS_OPCOES = ["SKU", "Quantidade solicitada"] + [
    campo
    for posicao in range(1, 5)
    for campo in [
        f"Código fornecedor {posicao}", f"Fornecedor {posicao}", f"Estoque {posicao}",
        f"Preço {posicao}", f"Variação custo {posicao}",
    ]
]


def _copiar_estilo_linha(ws, origem: int, destino: int, total_colunas: int) -> None:
    ws.row_dimensions[destino].height = ws.row_dimensions[origem].height
    for col in range(1, total_colunas + 1):
        cel_origem = ws.cell(origem, col)
        cel_destino = ws.cell(destino, col)
        if cel_origem.has_style:
            cel_destino._style = copy(cel_origem._style)
        if cel_origem.number_format:
            cel_destino.number_format = cel_origem.number_format
        if cel_origem.alignment:
            cel_destino.alignment = copy(cel_origem.alignment)
        if cel_origem.protection:
            cel_destino.protection = copy(cel_origem.protection)


def gerar_pedido_unificado(resultado: ResultadoMotor, template_path: str | Path) -> bytes:
    wb = load_workbook(template_path, keep_links=False, data_only=False)
    ws_base = wb["Base"]
    ws_opcoes = wb["Estoques_Fornecedores"]

    # Remove fórmulas e resíduos antigos apenas da área de dados do modelo.
    limite_base = max(355, 4 + len(resultado.pedido))
    for row in ws_base.iter_rows(min_row=5, max_row=limite_base, min_col=1, max_col=29):
        for cell in row:
            cell.value = None

    limite_opcoes = max(355, 2 + len(resultado.opcoes))
    for row in ws_opcoes.iter_rows(min_row=3, max_row=limite_opcoes, min_col=1, max_col=22):
        for cell in row:
            cell.value = None

    pedido = resultado.pedido.reindex(columns=COLUNAS_BASE)
    for idx, valores in enumerate(pedido.itertuples(index=False, name=None), start=5):
        if idx > 355:
            _copiar_estilo_linha(ws_base, 5, idx, 29)
        for col, valor in enumerate(valores, start=1):
            ws_base.cell(idx, col, None if pd.isna(valor) else valor)

    opcoes = resultado.opcoes.reindex(columns=COLUNAS_OPCOES)
    for idx, valores in enumerate(opcoes.itertuples(index=False, name=None), start=3):
        if idx > 355:
            _copiar_estilo_linha(ws_opcoes, 3, idx, 22)
        for col, valor in enumerate(valores, start=1):
            ws_opcoes.cell(idx, col, None if pd.isna(valor) else valor)

    ws_base.auto_filter.ref = f"A4:AC{max(4, 4 + len(pedido))}"
    ws_opcoes.auto_filter.ref = f"A2:V{max(2, 2 + len(opcoes))}"
    ws_base.freeze_panes = "A5"
    ws_opcoes.freeze_panes = "A3"

    # Garante formatos essenciais sem alterar o visual do template.
    for row in range(5, 5 + len(pedido)):
        for col in [15, 19]:
            ws_base.cell(row, col).number_format = "0.00%"
        ws_base.cell(row, 21).number_format = 'R$ #,##0.00'
    for row in range(3, 3 + len(opcoes)):
        for col in [6, 11, 16, 21]:
            ws_opcoes.cell(row, col).number_format = 'R$ #,##0.00'
        for col in [7, 12, 17, 22]:
            ws_opcoes.cell(row, col).number_format = "0.00%"

    wb.calculation.fullCalcOnLoad = False
    wb.calculation.forceFullCalc = False
    wb.calculation.calcMode = "manual"

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def dataframe_para_excel(abas: dict[str, pd.DataFrame]) -> bytes:
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="xlsxwriter") as writer:
        for nome, df in abas.items():
            nome_seguro = nome[:31]
            df.to_excel(writer, sheet_name=nome_seguro, index=False)
            ws = writer.sheets[nome_seguro]
            ws.freeze_panes(1, 0)
            ws.autofilter(0, 0, max(len(df), 1), max(len(df.columns) - 1, 0))
            for idx, coluna in enumerate(df.columns):
                largura = min(max(12, len(str(coluna)) + 2), 42)
                ws.set_column(idx, idx, largura)
    return buffer.getvalue()


def gerar_resumo_excel(resultado: ResultadoMotor) -> bytes:
    indicadores = pd.DataFrame(
        [
            ["ID da carga", resultado.id_carga],
            ["Aba da necessidade", resultado.diagnostico.get("aba_necessidade", "")],
            ["SKUs com Pedido Efetivo", resultado.diagnostico.get("skus_com_pedido", 0)],
            ["Unidades solicitadas", resultado.diagnostico.get("unidades_solicitadas", 0)],
            ["Valor total", resultado.resumo["valor_total"]],
            ["SKUs na necessidade", resultado.resumo["skus_necessidade"]],
            ["SKUs no pedido", resultado.resumo["skus_pedido"]],
            ["Pendências", resultado.resumo["pendencias"]],
            ["Fornecedores", resultado.resumo["fornecedores"]],
            ["Itens via busca ampliada", resultado.resumo["busca_ampliada"]],
        ],
        columns=["Indicador", "Valor"],
    )
    return dataframe_para_excel(
        {
            "Resumo": indicadores,
            "Por fornecedor": resultado.resumo["por_fornecedor"],
            "Motivos pendência": resultado.resumo["motivos_pendencia"],
        }
    )
