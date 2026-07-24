from __future__ import annotations

import unittest
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook

from src.exportacao import gerar_pedido_unificado
from src.motor import executar_motor
from src.leitura import ler_historico


class TestMotor(unittest.TestCase):
    def setUp(self) -> None:
        self.cadastro = pd.DataFrame(
            [
                ["111", "111", "1", "Produto A", "FAB A", "RX", 1, 1, "Ativo"],
                ["222", "222", "2", "Produto B", "FAB B", "MIP", 1, 1, "Ativo"],
                ["333", "334", "3", "Produto C", "FAB C", "SIMILAR", 10, 10, "Ativo"],
            ],
            columns=["ean_compra", "ean_venda", "sku", "descricao_oficial", "fabricante", "categoria", "caixaria_padrao", "multiplo_padrao", "status_ean"],
        )
        self.regras = pd.DataFrame(
            [
                ["Solfarma", "Distribuidor", "Sim", "Não", "Sim", "Sim", "28 dias", 2],
                ["Bloqueado", "Distribuidor", "Sim", "Sim", "Sim", "Não", "", 0],
            ],
            columns=["fornecedor", "tipo_operacao", "ativo", "bloqueado", "participa_cotacao", "participa_busca", "prazo_pagamento", "lead_time"],
        )
        self.necessidade = pd.DataFrame(
            [
                ["1", "111", "Produto A", 10, 10.5, "Não"],
                ["2", "222", "Produto B", 20, 5.5, "Não"],
                ["3", "334", "Produto C", 15, 21, "Não"],
            ],
            columns=["sku", "ean", "descricao", "quantidade_solicitada", "ultimo_custo", "ruptura_cronica"],
        )
        self.cotacao = pd.DataFrame(
            [
                ["C1", "Solfarma", "Distribuidor", "111", "Produto A", 10, "Unitário", 1, 1, 100],
                ["C2", "Bloqueado", "Distribuidor", "222", "Produto B", 4, "Unitário", 1, 1, 100],
                ["C3", "Solfarma", "Distribuidor", "222", "Produto B", 5, "Unitário", 1, 1, 100],
                ["C4", "Solfarma", "Distribuidor", "333 / 334", "Produto C", 200, "Caixa", 10, 10, 50],
            ],
            columns=["id_carga", "fornecedor", "tipo_operacao", "ean", "descricao_recebida", "preco_final", "tipo_preco", "embalagem", "multiplo", "estoque_fornecedor"],
        )

    def test_decisao_e_caixaria(self) -> None:
        resultado = executar_motor(self.cotacao, self.necessidade, self.cadastro, self.regras)
        por_sku = resultado.pedido.set_index("SKU")
        self.assertEqual(por_sku.loc["1", "Fornecedor recomendado"], "Solfarma")
        self.assertEqual(por_sku.loc["2", "Fornecedor recomendado"], "Solfarma")
        self.assertEqual(por_sku.loc["3", "Quantidade Solicitada"], 20)
        self.assertAlmostEqual(float(por_sku.loc["3", "Preço recomendado"]), 20.0)
        # A oferta bloqueada fica registrada no histórico, mas não vira pendência
        # quando existe outra opção válida para o SKU.
        self.assertNotIn("Fornecedor bloqueado", set(resultado.pendencias.get("Pendência", [])))
        self.assertTrue(resultado.id_carga.startswith("QDC_"))
        self.assertEqual(set(resultado.pedido["NR COTAÇÃO"]), {resultado.id_carga})


    def test_alias_fornecedor_razao_social(self) -> None:
        regras = pd.DataFrame(
            [["SOLFARMA COM PROD FARM SA", "Distribuidor", "Sim", "Não", "Sim", "Sim", "60", 3]],
            columns=["fornecedor", "tipo_operacao", "ativo", "bloqueado", "participa_cotacao", "participa_busca", "prazo_pagamento", "lead_time"],
        )
        cotacao = self.cotacao.iloc[[0]].copy()
        cotacao.loc[:, "fornecedor"] = "SOLFARMA"
        resultado = executar_motor(cotacao, self.necessidade.iloc[[0]], self.cadastro, regras)
        self.assertEqual(len(resultado.pedido), 1)
        self.assertEqual(resultado.pedido.iloc[0]["Fornecedor recomendado"], "SOLFARMA")
        self.assertEqual(resultado.diagnostico.get("fornecedores_sem_regra"), [])

    def test_leitor_mapa_diario_carrega_apenas_classificacao(self) -> None:
        caminho = Path("/tmp/mapa diario teste.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Planilha1"
        ws.append(["controle"])
        ws.append(["Código", "EAN", "", "Descrição", "OL/DIRETO/DIST", "VAN", "FABRICANTE ", "Cód Direto"])
        ws.append([1, 111, "", "Produto A", "OL", "VAN A", "FAB A", ""])
        ws.append([2, 222, "", "Produto B", "DIRETO", "VAN B", "FAB B", 999])
        wb.save(caminho)
        lido = ler_historico(caminho)
        self.assertEqual(lido.attrs.get("formato_historico"), "mapa_diario")
        self.assertFalse(lido.attrs.get("uso_busca_ampliada"))
        self.assertEqual(lido.set_index("sku").loc["1", "tipo_operacao"], "OL")
        self.assertEqual(lido.set_index("sku").loc["2", "tipo_operacao"], "DIRETO")

    def test_mapa_diario_classifica_por_sku(self) -> None:
        historico = pd.DataFrame(
            [
                ["1", "111", "OL"],
                ["2", "222", "DIRETO"],
                ["3", "334", "DIST"],
            ],
            columns=["sku", "ean", "tipo_operacao"],
        )
        historico.attrs["formato_historico"] = "mapa_diario"
        historico.attrs["uso_busca_ampliada"] = False

        cotacao = self.cotacao.copy()
        cotacao["tipo_operacao"] = ""
        resultado = executar_motor(
            cotacao, self.necessidade, self.cadastro, self.regras, historico=historico
        )
        tipos = resultado.pedido.set_index("SKU")["Tratativa"].to_dict()
        self.assertEqual(tipos["1"], "OL")
        self.assertEqual(tipos["2"], "Direto")
        self.assertEqual(tipos["3"], "Distribuidor")
        # O Mapa Diário é referência de classificação, não um histórico de
        # ofertas a ser anexado ao arquivo consolidado.
        self.assertNotIn("mapa_diario", set(resultado.historico.get("Origem do registro", [])))

    def test_template_mantem_estrutura(self) -> None:
        resultado = executar_motor(self.cotacao, self.necessidade, self.cadastro, self.regras)
        template = Path(__file__).resolve().parents[1] / "templates" / "Modelo Envio Pedidos Fornecedor_Medicamentos.xlsx"
        dados = gerar_pedido_unificado(resultado, template)
        saida = Path("/tmp/pedido_unificado_teste.xlsx")
        saida.write_bytes(dados)
        wb = load_workbook(saida, read_only=True, data_only=False)
        self.assertEqual(wb.sheetnames, ["Base", "Estoques_Fornecedores"])
        self.assertEqual(wb["Base"]["J4"].value, "SKU")
        self.assertEqual(wb["Estoques_Fornecedores"]["D2"].value, "Fornecedor 1")
        self.assertEqual(str(wb["Base"]["J5"].value), "1")
        self.assertEqual(wb["Estoques_Fornecedores"]["D3"].value, "Solfarma")
        wb.close()


if __name__ == "__main__":
    unittest.main()
